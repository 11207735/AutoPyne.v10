import os
import json
import sqlite3
import datetime

# Database directories and SQLite file paths
DB_DIR = "AGM-AGAFAY"
DB_SQLITE_FILE = os.path.join(DB_DIR, "agm_travel.db")
WORKSPACE_DB_DIR = "database"
WORKSPACE_SQLITE_FILE = os.path.join(WORKSPACE_DB_DIR, "agm_travel.db")
OLD_DB_FILE = os.path.join(DB_DIR, "database.json")
OLD_DATA_FILE = os.path.join(DB_DIR, "data.json")
OLD_WORKSPACE_DB_FILE = os.path.join("database", "agm_daily_trips.json")

def get_db_connection():
    """Returns a SQLite connection with Row factory enabled."""
    initialize_system()
    conn = sqlite3.connect(DB_SQLITE_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def levenshtein_distance(s1, s2):
    """Calculate Levenshtein distance for fuzzy name typo matching."""
    if len(s1) < len(s2):
        return levenshtein_distance(s2, s1)
    if len(s2) == 0:
        return len(s1)
    previous_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        current_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = previous_row[j + 1] + 1
            deletions = current_row[j] + 1
            substitutions = previous_row[j] + (c1 != c2)
            current_row.append(min(insertions, deletions, substitutions))
        previous_row = current_row
    return previous_row[-1]

def get_canonical_name(input_name, known_names):
    """Finds the canonical name using fuzzy typo matching."""
    cleaned = str(input_name).strip().upper()
    if not cleaned or cleaned in ["WITHOUT GUIDE", "?", "NONE", "UNASSIGNED"]:
        return cleaned

    # Exact match first
    for k in known_names:
        if cleaned == k.strip().upper():
            return k

    best_match = cleaned
    min_dist = 999

    for k in known_names:
        k_upper = k.strip().upper()
        if not k_upper or k_upper in ["WITHOUT GUIDE", "?", "NONE", "UNASSIGNED"]:
            continue
        dist = levenshtein_distance(cleaned, k_upper)
        max_allowed = 1 if len(cleaned) <= 4 else (2 if len(cleaned) <= 8 else 3)
        if dist <= max_allowed and dist < min_dist:
            min_dist = dist
            best_match = k

    return best_match

def parse_date_parts(date_str):
    """
    Robustly parses standard dates in DD-MM-YYYY or DD/MM/YYYY formats.
    Returns strings (day, month, year) with proper zero-padding.
    """
    date_str = str(date_str).strip()
    for sep in ["-", "/"]:
        if sep in date_str:
            parts = date_str.split(sep)
            if len(parts) >= 3:
                d = parts[0].strip()
                m = parts[1].strip()
                y = parts[2].strip()
                if len(y) == 2 and y.isdigit():
                    y = "20" + y
                if len(d) == 1 and d.isdigit():
                    d = "0" + d
                if len(m) == 1 and m.isdigit():
                    m = "0" + m
                return d, m, y
    
    # Fallback to current date
    today = datetime.datetime.now()
    return today.strftime("%d"), today.strftime("%m"), today.strftime("%Y")

def initialize_system():
    """Create the SQLite database and create necessary tables and indexes with auto-repair for corrupt databases."""
    if not os.path.exists(DB_DIR):
        try:
            os.makedirs(DB_DIR)
        except Exception:
            pass
    if not os.path.exists(WORKSPACE_DB_DIR):
        try:
            os.makedirs(WORKSPACE_DB_DIR)
        except Exception:
            pass
    
    # Ensure Desktop/AGM-AGAFAY, Documents/AGM-AGAFAY and AGM-WorkSpace exist
    home = os.path.expanduser("~")
    for parent in ["Desktop", "Documents"]:
        for f_name in ["AGM-AGAFAY", "AGM-WorkSpace"]:
            folder_path = os.path.join(home, parent, f_name)
            try:
                if not os.path.exists(folder_path):
                    os.makedirs(folder_path)
            except Exception:
                pass

    try:
        # Initialize SQLite Database & Schema
        conn = sqlite3.connect(DB_SQLITE_FILE)
        cur = conn.cursor()

        # 1. Trips Table
        cur.execute("""
        CREATE TABLE IF NOT EXISTS trips (
            id INTEGER PRIMARY KEY,
            van_type TEXT NOT NULL DEFAULT 'Big van',
            guide TEXT NOT NULL DEFAULT '',
            driver TEXT NOT NULL DEFAULT '',
            company TEXT NOT NULL DEFAULT 'AGM',
            pax TEXT NOT NULL DEFAULT '0',
            quads TEXT NOT NULL DEFAULT '0',
            camels TEXT NOT NULL DEFAULT '0',
            person_extra TEXT DEFAULT 'None',
            quad_extra TEXT DEFAULT 'None',
            camel_extra TEXT DEFAULT 'None',
            extra_payment TEXT DEFAULT '0 DH',
            date TEXT NOT NULL,
            time TEXT NOT NULL,
            drivers_list TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );
        """)

        # Indexes for fast querying
        cur.execute("CREATE INDEX IF NOT EXISTS idx_trips_date ON trips(date);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_trips_guide ON trips(guide);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_trips_driver ON trips(driver);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_trips_company ON trips(company);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_trips_date_time ON trips(date, time);")

        # 2. Registered Guides Table
        cur.execute("""
        CREATE TABLE IF NOT EXISTS registered_guides (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            nickname TEXT,
            phone TEXT,
            dates_worked TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now'))
        );
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_guides_name ON registered_guides(name);")

        # 3. Registered Drivers Table
        cur.execute("""
        CREATE TABLE IF NOT EXISTS registered_drivers (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            van_type TEXT DEFAULT 'Big van',
            company_name TEXT DEFAULT 'AGM',
            phone TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now'))
        );
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_drivers_name ON registered_drivers(name);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_drivers_company ON registered_drivers(company_name);")

        # 4. Managers Table
        cur.execute("""
        CREATE TABLE IF NOT EXISTS managers (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            lastname TEXT,
            school_level TEXT,
            skill TEXT,
            started_from TEXT,
            email TEXT,
            employee_id TEXT,
            pin TEXT,
            status TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );
        """)

        # 5. Payment Rates Table
        cur.execute("""
        CREATE TABLE IF NOT EXISTS payment_rates (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        """)

        # 6. Settled Payments Table
        cur.execute("""
        CREATE TABLE IF NOT EXISTS settled_payments (
            id TEXT PRIMARY KEY,
            entity_type TEXT NOT NULL,
            entity_name TEXT NOT NULL,
            period_key TEXT NOT NULL,
            amount_dh REAL NOT NULL,
            trips_count INTEGER DEFAULT 0,
            settled_at TEXT NOT NULL,
            settled_by TEXT NOT NULL,
            receipt_number TEXT NOT NULL,
            notes TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_settled_entity ON settled_payments(entity_type, entity_name);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_settled_period ON settled_payments(period_key);")

        conn.commit()

        # Check and migrate legacy database.json if present
        migrate_legacy_json_to_sqlite(conn)

        conn.close()
    except (sqlite3.DatabaseError, sqlite3.OperationalError) as err:
        print(f"[SQLite Warning] Corrupt or malformed database detected ({err}). Recreating database file...")
        try:
            if os.path.exists(DB_SQLITE_FILE):
                os.remove(DB_SQLITE_FILE)
        except Exception:
            pass
        # Retry once with a fresh database
        conn = sqlite3.connect(DB_SQLITE_FILE)
        cur = conn.cursor()
        cur.execute("""
        CREATE TABLE IF NOT EXISTS trips (
            id INTEGER PRIMARY KEY, van_type TEXT NOT NULL DEFAULT 'Big van', guide TEXT NOT NULL DEFAULT '',
            driver TEXT NOT NULL DEFAULT '', company TEXT NOT NULL DEFAULT 'AGM', pax TEXT NOT NULL DEFAULT '0',
            quads TEXT NOT NULL DEFAULT '0', camels TEXT NOT NULL DEFAULT '0', person_extra TEXT DEFAULT 'None',
            quad_extra TEXT DEFAULT 'None', camel_extra TEXT DEFAULT 'None', extra_payment TEXT DEFAULT '0 DH',
            date TEXT NOT NULL, time TEXT NOT NULL, drivers_list TEXT, created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_trips_date ON trips(date);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_trips_guide ON trips(guide);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_trips_driver ON trips(driver);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_trips_company ON trips(company);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_trips_date_time ON trips(date, time);")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS registered_guides (
            id TEXT PRIMARY KEY, name TEXT NOT NULL, nickname TEXT, phone TEXT, dates_worked TEXT,
            is_active INTEGER DEFAULT 1, created_at TEXT DEFAULT (datetime('now'))
        );
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS registered_drivers (
            id TEXT PRIMARY KEY, name TEXT NOT NULL, van_type TEXT DEFAULT 'Big van',
            company_name TEXT DEFAULT 'AGM', phone TEXT, is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now'))
        );
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS managers (
            id TEXT PRIMARY KEY, name TEXT NOT NULL, lastname TEXT, school_level TEXT, skill TEXT,
            started_from TEXT, email TEXT, employee_id TEXT, pin TEXT, status TEXT, created_at TEXT DEFAULT (datetime('now'))
        );
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS payment_rates (key TEXT PRIMARY KEY, value TEXT NOT NULL);
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS settled_payments (
            id TEXT PRIMARY KEY, entity_type TEXT NOT NULL, entity_name TEXT NOT NULL, period_key TEXT NOT NULL,
            amount_dh REAL NOT NULL, trips_count INTEGER DEFAULT 0, settled_at TEXT NOT NULL, settled_by TEXT NOT NULL,
            receipt_number TEXT NOT NULL, notes TEXT, created_at TEXT DEFAULT (datetime('now'))
        );
        """)
        conn.commit()
        conn.close()

def migrate_legacy_json_to_sqlite(conn):
    """Auto-migrates any old data.json or database.json records into the SQLite database and deletes the JSON files."""
    legacy_records = []
    
    # 1. Check DB_FILE and OLD_DATA_FILE
    for old_path in [OLD_DB_FILE, OLD_DATA_FILE, "data.json", "database.json"]:
        if os.path.exists(old_path):
            try:
                with open(old_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        for item in data:
                            if not any(r.get("id") == item.get("id") for r in legacy_records):
                                legacy_records.append(item)
            except Exception:
                pass

    # 2. Check WORKSPACE_DB_FILE
    if os.path.exists(OLD_WORKSPACE_DB_FILE):
        try:
            with open(OLD_WORKSPACE_DB_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, list):
                    for item in data:
                        if not any(r.get("id") == item.get("id") for r in legacy_records):
                            legacy_records.append(item)
        except Exception:
            pass

    if legacy_records:
        cur = conn.cursor()
        for r in legacy_records:
            r_id = r.get("id")
            if r_id is None:
                continue
            cur.execute("""
            INSERT OR REPLACE INTO trips (
                id, van_type, guide, driver, company, pax, quads, camels,
                person_extra, quad_extra, camel_extra, extra_payment,
                date, time, drivers_list, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            """, (
                int(r_id),
                str(r.get("van_type", "Big van")),
                str(r.get("guide", "")),
                str(r.get("driver", "")),
                str(r.get("company", "AGM")),
                str(r.get("pax", "0")),
                str(r.get("quads", "0")),
                str(r.get("camels", "0")),
                str(r.get("person_extra", "None")),
                str(r.get("quad_extra", "None")),
                str(r.get("camel_extra", "None")),
                str(r.get("extra_payment", "0 DH")),
                str(r.get("date", "")),
                str(r.get("time", "")),
                json.dumps(r.get("driversList")) if r.get("driversList") else None
            ))
        conn.commit()

    # Safely delete legacy JSON files to complete migration and optimize disk speed
    for old_path in [OLD_DB_FILE, OLD_DATA_FILE, "data.json", "database.json", OLD_WORKSPACE_DB_FILE]:
        try:
            if os.path.exists(old_path):
                os.remove(old_path)
                print(f"[SQLite Optimization] Successfully migrated and removed legacy {old_path}")
        except Exception as e:
            pass

def trip_row_to_dict(row):
    """Converts a SQLite trip row to dictionary format matching the UI model."""
    d = dict(row)
    drivers_list = None
    if d.get("drivers_list"):
        try:
            drivers_list = json.loads(d["drivers_list"])
        except Exception:
            drivers_list = None
    res = {
        "id": d["id"],
        "van_type": d["van_type"],
        "guide": d["guide"],
        "driver": d["driver"],
        "company": d["company"],
        "pax": d["pax"],
        "quads": d["quads"],
        "camels": d["camels"],
        "person_extra": d["person_extra"],
        "quad_extra": d["quad_extra"],
        "camel_extra": d["camel_extra"],
        "extra_payment": d["extra_payment"],
        "date": d["date"],
        "time": d["time"]
    }
    if drivers_list:
        res["driversList"] = drivers_list
    return res

def load_records(limit=None, offset=None, search=None, date_filter=None):
    """Load records from SQLite database with optional pagination (LIMIT/OFFSET) and search filters."""
    conn = get_db_connection()
    cur = conn.cursor()

    query = "SELECT * FROM trips WHERE 1=1"
    params = []

    if date_filter:
        query += " AND date = ?"
        params.append(date_filter)

    if search:
        search_pattern = f"%{search}%"
        query += " AND (guide LIKE ? OR driver LIKE ? OR company LIKE ? OR pax LIKE ? OR date LIKE ?)"
        params.extend([search_pattern, search_pattern, search_pattern, search_pattern, search_pattern])

    query += " ORDER BY id DESC"

    if limit is not None:
        query += " LIMIT ?"
        params.append(int(limit))
        if offset is not None:
            query += " OFFSET ?"
            params.append(int(offset))

    cur.execute(query, params)
    rows = cur.fetchall()
    records = [trip_row_to_dict(r) for r in rows]
    conn.close()
    return records

def save_raw_records(records):
    """Save records directly into SQLite database (bulk upsert) and clean up legacy JSON files."""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM trips")
        for r in records:
            drivers_list_str = json.dumps(r.get("driversList")) if r.get("driversList") else None
            cur.execute("""
            INSERT OR REPLACE INTO trips (
                id, van_type, guide, driver, company, pax, quads, camels,
                person_extra, quad_extra, camel_extra, extra_payment,
                date, time, drivers_list, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            """, (
                r.get("id"),
                str(r.get("van_type", "Big van")),
                str(r.get("guide", "")),
                str(r.get("driver", "")),
                str(r.get("company", "AGM")),
                str(r.get("pax", "0")),
                str(r.get("quads", "0")),
                str(r.get("camels", "0")),
                str(r.get("person_extra", "None")),
                str(r.get("quad_extra", "None")),
                str(r.get("camel_extra", "None")),
                str(r.get("extra_payment", "0 DH")),
                str(r.get("date", "")),
                str(r.get("time", "")),
                drivers_list_str
            ))
        conn.commit()
    except Exception as e:
        print(f"[SQLite Save Error] {e}")
    finally:
        conn.close()

    # Safely clean up any legacy JSON files so system stays fast
    for old_path in [OLD_DB_FILE, OLD_DATA_FILE, "data.json", "database.json", OLD_WORKSPACE_DB_FILE]:
        try:
            if os.path.exists(old_path):
                os.remove(old_path)
        except Exception:
            pass

def is_record_complete(record):
    """
    Validation Algorithm:
    Checks if a record is complete.
    """
    guide = str(record.get("guide", "")).strip().upper()
    driver = str(record.get("driver", "")).strip()
    date_val = str(record.get("date", "")).strip()
    time_val = str(record.get("time", "")).strip()

    if not driver or driver in ["?", "NONE", "UNASSIGNED"]:
        return False
    if not date_val or not time_val:
        return False

    pax = str(record.get("pax", "")).strip()
    quads = str(record.get("quads", "")).strip()
    camels = str(record.get("camels", "")).strip()

    if pax == "?" or quads == "?" or camels == "?":
        return False

    return True

def get_stats():
    """Retrieve statistics for UI status bar."""
    records = load_records()
    done_count = sum(1 for r in records if is_record_complete(r))
    rest_count = len(records) - done_count
    return done_count, rest_count

def style_coords_range(ws, start_row, start_col, end_row, end_col, font=None, fill=None, border=None, alignment=None):
    """Utility function to apply styles to a contiguous cell range."""
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            if font is not None:
                cell.font = font
            if fill is not None:
                cell.fill = fill
            if border is not None:
                cell.border = border
            if alignment is not None:
                cell.alignment = alignment

def sync_to_excel():
    """
    Export and rebuild corporate executive Excel workbooks grouped by year and month.
    Designed in the style of an elite Holding Company master operations workbook.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[Excel Sync] openpyxl not installed. Skipping direct Excel binary generation.")
        return

    records = load_records()
    
    today_now = datetime.datetime.now()
    today_day = today_now.strftime("%d")
    today_month = today_now.strftime("%m")
    today_year = today_now.strftime("%Y")
    today_str = f"{today_day}-{today_month}-{today_year}"

    # Group records by (year, month)
    monthly_data = {}
    monthly_data[(today_year, today_month)] = []

    for r in records:
        date_val = r.get("date", "")
        _, m, y = parse_date_parts(date_val)
        key = (y, m)
        if key not in monthly_data:
            monthly_data[key] = []
        monthly_data[key].append(r)

    # Rebuild Excel file for each specific month & year
    for (year, month), items in monthly_data.items():
        wb = Workbook()
        ws = wb.active
        ws.title = f"AGM-OPERATIONS {month}-{year}"
        
        # Keep crisp grid lines
        ws.views.sheetView[0].showGridLines = True

        def sort_key(x):
            d_val, m_val, y_val = parse_date_parts(x.get("date", ""))
            t_parts = str(x.get("time", "00:00")).split(":")
            day_num = int(d_val) if d_val.isdigit() else 1
            hour = int(t_parts[0]) if len(t_parts) >= 1 and t_parts[0].isdigit() else 0
            minute = int(t_parts[1]) if len(t_parts) >= 2 and t_parts[1].isdigit() else 0
            return (day_num, hour, minute)
            
        items_sorted = sorted(items, key=sort_key)

        daily_groups = {}
        for item in items_sorted:
            day_str = item.get("date", "Unknown-Date")
            if day_str not in daily_groups:
                daily_groups[day_str] = []
            daily_groups[day_str].append(item)

        if year == today_year and month == today_month:
            if today_str not in daily_groups:
                daily_groups[today_str] = []

        # ================= DESIGN PALETTE: BIG HOLDING COMPANY MASTER THEME =================
        # Fonts
        corp_title_font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
        corp_sub_font = Font(name="Calibri", size=9.5, italic=True, color="D1D5DB")
        day_banner_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=10, color="1F2937")
        data_font_bold = Font(name="Calibri", size=10, bold=True, color="111827")
        status_done_font = Font(name="Calibri", size=9.5, bold=True, color="065F46")
        status_pending_font = Font(name="Calibri", size=9.5, bold=True, color="92400E")
        totals_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
        
        # Fills (Holding Corporate Palette)
        corp_banner_fill = PatternFill(start_color="0F2537", end_color="0F2537", fill_type="solid") # Deep Navy Executive
        corp_sub_fill = PatternFill(start_color="1B3B54", end_color="1B3B54", fill_type="solid")    # Slate Navy
        day_banner_fill = PatternFill(start_color="234966", end_color="234966", fill_type="solid")  # Corporate Blue
        header_fill = PatternFill(start_color="142836", end_color="142836", fill_type="solid")      # Table Header Navy
        
        row_even_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")    # Clean White
        row_odd_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")     # Subtle Ice Tint
        
        status_done_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Soft Emerald Mint
        status_pending_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Soft Amber Bronze
        
        totals_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")     # Platinum Steel Gray
        
        # Alignments
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right_align = Alignment(horizontal="right", vertical="center")
        
        # Borders
        thin_slate = Side(border_style="thin", color="CBD5E1")
        thin_steel = Side(border_style="thin", color="94A3B8")
        double_navy = Side(border_style="double", color="0F2537")
        thick_navy = Side(border_style="medium", color="0F2537")
        
        cell_border = Border(left=thin_slate, right=thin_slate, top=thin_slate, bottom=thin_slate)
        header_border = Border(left=thin_steel, right=thin_steel, top=thin_steel, bottom=thick_navy)
        total_row_border = Border(top=thin_steel, bottom=double_navy, left=thin_slate, right=thin_slate)

        # Table Columns: Column 3 (C) to Column 14 (N)
        start_col = 3
        end_col = 14
        current_row = 2

        # 1. HOLDING COMPANY MASTER HEADER BANNER
        ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
        ws.row_dimensions[current_row].height = 36
        style_coords_range(ws, current_row, start_col, current_row, end_col, font=corp_title_font, fill=corp_banner_fill, alignment=center_align)
        ws.cell(row=current_row, column=start_col, value="AGM GROUP HOLDING • OPERATIONS & DISPATCH MASTER WORKBOOK")
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
        ws.row_dimensions[current_row].height = 22
        style_coords_range(ws, current_row, start_col, current_row, end_col, font=corp_sub_font, fill=corp_sub_fill, alignment=center_align)
        ws.cell(row=current_row, column=start_col, value=f"OPERATIONAL DIVISION: AGAFAY EXPEDITIONS | PERIOD: {month}/{year} | MASTER SYSTEM RECONCILIATION")
        current_row += 2

        def day_sort_key(day_val):
            d_val, m_val, y_val = parse_date_parts(day_val)
            try:
                return int(d_val)
            except ValueError:
                return 1

        sorted_days = sorted(daily_groups.keys(), key=day_sort_key)

        for day_str in sorted_days:
            day_items = daily_groups[day_str]
            formatted_title_date = day_str.replace("-", "/")

            # Day Section Header Banner
            ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
            ws.row_dimensions[current_row].height = 30
            style_coords_range(ws, current_row, start_col, current_row, end_col, font=day_banner_font, fill=day_banner_fill, alignment=left_align)
            ws.cell(row=current_row, column=start_col, value=f"   WORKDAY OPERATIONS LOG: {formatted_title_date} (TOTAL TRIPS: {len(day_items)})")
            current_row += 1

            # Column Headers
            headers = [
                "NO",
                "GUIDE NAME",
                "DRIVER NAME",
                "VAN TYPE",
                "COMPANY",
                "PAX",
                "QUADS",
                "CAMELS",
                "EXTRA DETAILS",
                "EXTRA REV (DH)",
                "DATE",
                "TIME"
            ]
            ws.row_dimensions[current_row].height = 26
            for col_idx, h in enumerate(headers, start_col):
                cell = ws.cell(row=current_row, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = header_border
            current_row += 1

            total_pax = 0
            total_quads = 0
            total_camels = 0
            total_extra_rev = 0

            if not day_items:
                ws.row_dimensions[current_row].height = 22
                for col_idx in range(start_col, end_col + 1):
                    cell = ws.cell(row=current_row, column=col_idx, value="—")
                    cell.font = data_font
                    cell.fill = row_even_fill
                    cell.alignment = center_align
                    cell.border = cell_border
                current_row += 1
            else:
                for idx, item in enumerate(day_items):
                    is_even = (idx % 2 == 0)
                    row_fill = row_even_fill if is_even else row_odd_fill

                    ws.row_dimensions[current_row].height = 23

                    g_val = str(item.get("guide", "WITHOUT GUIDE")).upper()
                    d_val = str(item.get("driver", "UNASSIGNED")).upper()
                    v_val = str(item.get("van_type", "Big van"))
                    comp_val = str(item.get("company", "AGM")).upper()
                    p_val = str(item.get("pax", "?"))
                    q_val = str(item.get("quads", "?"))
                    c_val = str(item.get("camels", "?"))
                    date_val = str(item.get("date", ""))
                    time_val = str(item.get("time", ""))

                    # Extras calculation
                    extras_list = []
                    p_extra = str(item.get("person_extra", "None"))
                    q_extra = str(item.get("quad_extra", "None"))
                    c_extra = str(item.get("camel_extra", "None"))

                    if p_extra not in ["None", "0", ""]:
                        extras_list.append(f"Pax: {p_extra}")
                    if q_extra not in ["None", "0", ""]:
                        extras_list.append(f"Quads: {q_extra}")
                    if c_extra not in ["None", "0", ""]:
                        extras_list.append(f"Camels: {c_extra}")

                    extras_str = " | ".join(extras_list) if extras_list else "None"
                    
                    # Extra payment extraction
                    extra_pay_raw = str(item.get("extra_payment", "0 DH"))
                    extra_pay_num = 0
                    try:
                        extra_pay_num = int("".join(filter(str.isdigit, extra_pay_raw))) if any(c.isdigit() for c in extra_pay_raw) else 0
                    except ValueError:
                        extra_pay_num = 0

                    try:
                        total_pax += int(p_val)
                    except ValueError:
                        pass
                    try:
                        total_quads += int(q_val)
                    except ValueError:
                        pass
                    try:
                        total_camels += int(c_val)
                    except ValueError:
                        pass
                    total_extra_rev += extra_pay_num

                    row_values = [
                        idx + 1,
                        g_val,
                        d_val,
                        v_val,
                        comp_val,
                        int(p_val) if p_val.isdigit() else p_val,
                        int(q_val) if q_val.isdigit() else q_val,
                        int(c_val) if c_val.isdigit() else c_val,
                        extras_str,
                        f"{extra_pay_num} DH" if extra_pay_num > 0 else "0 DH",
                        date_val,
                        time_val
                    ]

                    for col_offset, val in enumerate(row_values):
                        col_num = start_col + col_offset
                        cell = ws.cell(row=current_row, column=col_num, value=val)
                        cell.font = data_font_bold if col_offset in [0, 1, 2] else data_font
                        cell.fill = row_fill
                        cell.alignment = left_align if col_offset in [1, 2, 8] else center_align
                        cell.border = cell_border

                    current_row += 1

            # Day Totals Row
            ws.row_dimensions[current_row].height = 26
            ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=start_col + 4)
            style_coords_range(ws, current_row, start_col, current_row, end_col, font=totals_font, fill=totals_fill, border=total_row_border, alignment=center_align)
            
            ws.cell(row=current_row, column=start_col, value="DAILY RECONCILED TOTALS")
            ws.cell(row=current_row, column=start_col + 5, value=total_pax)
            ws.cell(row=current_row, column=start_col + 6, value=total_quads)
            ws.cell(row=current_row, column=start_col + 7, value=total_camels)
            ws.cell(row=current_row, column=start_col + 8, value="TOTAL EXTRA REVENUE:")
            ws.cell(row=current_row, column=start_col + 9, value=f"{total_extra_rev} DH")

            current_row += 3

        # ================= MONTHLY SUMMARY ANALYTICS TABLES =================
        # 1. Guides Summary
        known_guides = []
        for r in items:
            g = str(r.get("guide", "")).strip().upper()
            if g and g not in ["WITHOUT GUIDE", "?", "NONE", "UNASSIGNED"]:
                canon = get_canonical_name(g, known_guides)
                if canon == g and g not in known_guides:
                    known_guides.append(g)

        guide_dates = {}
        for r in items:
            g = str(r.get("guide", "")).strip().upper()
            d_str = str(r.get("date", "")).strip()
            if g and g not in ["WITHOUT GUIDE", "?", "NONE", "UNASSIGNED"]:
                canon = get_canonical_name(g, known_guides)
                if canon not in guide_dates:
                    guide_dates[canon] = set()
                if d_str:
                    guide_dates[canon].add(d_str)

        ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=start_col + 3)
        ws.row_dimensions[current_row].height = 28
        style_coords_range(ws, current_row, start_col, current_row, start_col + 3, font=day_banner_font, fill=day_banner_fill, alignment=left_align)
        ws.cell(row=current_row, column=start_col, value="   MONTHLY GUIDES ATTENDANCE & SUMMARY")
        current_row += 1

        guide_headers = ["Guide Name", "Days Worked", "Total Shifts", "Dates Active"]
        ws.row_dimensions[current_row].height = 24
        for idx, gh in enumerate(guide_headers):
            cell = ws.cell(row=current_row, column=start_col + idx, value=gh)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = header_border
        current_row += 1

        for g_name in sorted(guide_dates.keys()):
            dates_list = sorted(list(guide_dates[g_name]), key=day_sort_key)
            days_worked = len(dates_list)
            dates_str = ", ".join(dates_list)
            ws.row_dimensions[current_row].height = 22

            c1 = ws.cell(row=current_row, column=start_col, value=g_name)
            c2 = ws.cell(row=current_row, column=start_col + 1, value=days_worked)
            c3 = ws.cell(row=current_row, column=start_col + 2, value=days_worked)
            c4 = ws.cell(row=current_row, column=start_col + 3, value=dates_str)

            for c_idx, cell in enumerate([c1, c2, c3, c4]):
                cell.font = data_font
                cell.fill = row_even_fill
                cell.alignment = left_align if c_idx in [0, 3] else center_align
                cell.border = cell_border
            current_row += 1

        # Column Widths auto-tuning
        for col_idx in range(start_col, end_col + 1):
            col_letter = get_column_letter(col_idx)
            if col_idx in [start_col + 1, start_col + 2]:  # Guide and Driver
                ws.column_dimensions[col_letter].width = 24
            elif col_idx in [start_col + 8]:  # Extra details
                ws.column_dimensions[col_letter].width = 22
            else:
                ws.column_dimensions[col_letter].width = 15

        # Save workbook to Desktop, Documents, and local AGM-AGAFAY
        home = os.path.expanduser("~")
        dest_folders = [
            os.path.join(home, "Desktop", "AGM-AGAFAY"),
            os.path.join(home, "Documents", "AGM-AGAFAY"),
            os.path.join(home, "Desktop", "AGM-WorkSpace", "Years", f"year-{year}", f"{month}-{year}"),
            DB_DIR
        ]
        for base_folder in dest_folders:
            try:
                if not os.path.exists(base_folder):
                    os.makedirs(base_folder)
                file_path = os.path.join(base_folder, f"{month}-{year}.xlsx")
                wb.save(file_path)
            except Exception:
                pass

def add_or_update_record(record_id, van_type, guide, driver, pax, quads, camels, date_str, time_str, company="AGM", person_extra="None", quad_extra="None", camel_extra="None", extra_payment="0 DH", drivers_list=None):
    """
    Insert a new record or update an existing one in the SQLite database via parameterized SQL,
    then automatically sync the changes to rebuild the Excel sheets.
    """
    conn = get_db_connection()
    cur = conn.cursor()
    
    van_type = (van_type or "Big van").strip()
    driver = (driver or "").strip().upper()
    pax = (pax or "?").strip()
    quads = (quads or "?").strip()
    camels = (camels or "?").strip()
    date_str = (date_str or "").strip()
    time_str = (time_str or "").strip()
    company = (company or "AGM").strip().upper()

    if van_type.lower() == "mini van" and (not guide or guide.upper() in ["NONE", "H1", "WITHOUT GUIDE"]):
        guide = "WITHOUT GUIDE"
    else:
        guide = (guide or "").strip().upper()

    drivers_json_str = json.dumps(drivers_list) if drivers_list else None

    if record_id is not None:
        target_id = int(record_id)
        cur.execute("""
        UPDATE trips SET
            van_type = ?,
            guide = ?,
            driver = ?,
            company = ?,
            pax = ?,
            quads = ?,
            camels = ?,
            person_extra = ?,
            quad_extra = ?,
            camel_extra = ?,
            extra_payment = ?,
            date = ?,
            time = ?,
            drivers_list = ?,
            updated_at = datetime('now')
        WHERE id = ?;
        """, (
            van_type, guide, driver, company, pax, quads, camels,
            person_extra, quad_extra, camel_extra, extra_payment,
            date_str, time_str, drivers_json_str, target_id
        ))
        if cur.rowcount == 0:
            cur.execute("""
            INSERT INTO trips (
                id, van_type, guide, driver, company, pax, quads, camels,
                person_extra, quad_extra, camel_extra, extra_payment,
                date, time, drivers_list, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'));
            """, (
                target_id, van_type, guide, driver, company, pax, quads, camels,
                person_extra, quad_extra, camel_extra, extra_payment,
                date_str, time_str, drivers_json_str
            ))
    else:
        cur.execute("""
        INSERT INTO trips (
            van_type, guide, driver, company, pax, quads, camels,
            person_extra, quad_extra, camel_extra, extra_payment,
            date, time, drivers_list, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'));
        """, (
            van_type, guide, driver, company, pax, quads, camels,
            person_extra, quad_extra, camel_extra, extra_payment,
            date_str, time_str, drivers_json_str
        ))

    conn.commit()
    conn.close()
    sync_to_excel()

def delete_record(record_id):
    """Delete a record from SQLite database using SQL DELETE and update Excel."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM trips WHERE id = ?;", (int(record_id),))
    conn.commit()
    conn.close()
    sync_to_excel()

if __name__ == "__main__":
    import argparse
    import sys

    parser = argparse.ArgumentParser(description="AGM Travel Enterprise SQLite Backend Manager")
    parser.add_argument("--action", type=str, required=True, choices=["init", "load", "add_or_update", "delete", "sync", "check_today"])
    parser.add_argument("--id", type=str, default=None)
    parser.add_argument("--van_type", type=str, default="Big van")
    parser.add_argument("--guide", type=str, default="")
    parser.add_argument("--driver", type=str, default="")
    parser.add_argument("--company", type=str, default="AGM")
    parser.add_argument("--pax", type=str, default="?")
    parser.add_argument("--quads", type=str, default="?")
    parser.add_argument("--camels", type=str, default="?")
    parser.add_argument("--person_extra", type=str, default="None")
    parser.add_argument("--quad_extra", type=str, default="None")
    parser.add_argument("--camel_extra", type=str, default="None")
    parser.add_argument("--extra_payment", type=str, default="0 DH")
    parser.add_argument("--date", type=str, default="")
    parser.add_argument("--time", type=str, default="")
    parser.add_argument("--drivers_json", type=str, default=None)
    parser.add_argument("--page", type=int, default=None)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--search", type=str, default=None)

    args = parser.parse_args()

    try:
        if args.action == "init":
            initialize_system()
            print(json.dumps({"status": "success", "message": "AGM Travel SQLite Database initialized"}))
        elif args.action == "load":
            offset = None
            if args.page and args.limit:
                offset = (args.page - 1) * args.limit
            records = load_records(limit=args.limit, offset=offset, search=args.search, date_filter=args.date or None)
            print(json.dumps({"status": "success", "records": records, "count": len(records)}))
        elif args.action == "add_or_update":
            record_id = int(args.id) if args.id is not None and args.id != "null" and args.id != "" else None
            d_list = None
            if args.drivers_json:
                try:
                    d_list = json.loads(args.drivers_json)
                except Exception:
                    pass
            add_or_update_record(
                record_id,
                args.van_type,
                args.guide,
                args.driver,
                args.pax,
                args.quads,
                args.camels,
                args.date,
                args.time,
                args.company,
                args.person_extra,
                args.quad_extra,
                args.camel_extra,
                args.extra_payment,
                d_list
            )
            print(json.dumps({"status": "success", "message": "Record saved to SQLite and synced to Excel"}))
        elif args.action == "delete":
            if args.id is None:
                print(json.dumps({"status": "error", "message": "Missing record ID"}))
                sys.exit(1)
            delete_record(int(args.id))
            print(json.dumps({"status": "success", "message": "Record deleted from SQLite and synced"}))
        elif args.action == "sync":
            sync_to_excel()
            print(json.dumps({"status": "success", "message": "Excel files synchronized with Holding Company template"}))
        elif args.action == "check_today":
            initialize_system()
            sync_to_excel()
            today_str = datetime.datetime.now().strftime("%d-%m-%Y")
            print(json.dumps({"status": "success", "today": today_str}))
    except Exception as e:
        print(json.dumps({"status": "error", "message": str(e)}))
        sys.exit(1)
